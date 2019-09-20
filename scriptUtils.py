import openpyxl

### import openpyxl- help me to read XML file  ###

def getRowCount(file, fileName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(fileName)
    return (sheet.max_row)

def getColums(file,fileName):
    workbook = openpyxl.load_workbook(file)
    sheet  = workbook.get_sheet_by_name(fileName)
    return (sheet.max_column)

def readData (file,fileName,rownum,columno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(fileName)
    return sheet.cell(row=rownum, column=columno).value

def storeData (file,fileName,rownum,columno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(fileName)
    sheet.cell(row=rownum, column=columno).value =data
    workbook.save(file)







